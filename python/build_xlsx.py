#!/usr/bin/env python3
"""loads.json -> panel schedule workbook (schedule sheet + cable sizing sheet).

Usage:  python3 build_xlsx.py loads.json [outdir]
"""
import sys, os, copy
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from panel_core import load_config, build, safe_name, pe_size, fmt_mm2

TA = 'Tahoma'
N1 = '_(* #,##0.0_);_(* \\(#,##0.0\\);_(* \\-?_);_(@_)'
N2 = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)'
thin = Side(style='thin')
BOX = Border(thin, thin, thin, thin)
HDR = PatternFill('solid', fgColor='D9D9D9')
INP = PatternFill('solid', fgColor='FFF2CC')


def put(ws, r, c, v, bold=False, sz=9, fill=None, al='center', fmt=None, border=True, wrap=False):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=TA, size=sz, bold=bold)
    cell.alignment = Alignment(horizontal=al, vertical='center', wrap_text=wrap)
    if border:
        cell.border = BOX
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def sheet_schedule(wb, m):
    p, s = m['panel'], m['sys']
    ws = wb.create_sheet((p.get('name', 'PANEL'))[:31])
    ws['A1'] = f"{p.get('name','PANEL')}   {p.get('reference','')}".strip()
    ws['A1'].font = Font(name=TA, size=11, bold=True)
    ws['E2'] = 'load (watt)'
    ws['E2'].font = Font(name=TA, size=10)
    ws.merge_cells('E2:G2')
    for c, t in ((1, 'No'), (2, 'Description'), (4, 'PHASE'), (5, 'R'), (6, 'S'), (7, 'T'), (8, 'Amphere')):
        put(ws, 3, c, t, sz=10)
        ws.merge_cells(start_row=3, start_column=c, end_row=6, end_column=c)
    put(ws, 4, 3, 'LOAD', sz=7)
    put(ws, 7, 2, p.get('group', 'HVAC'), bold=True, sz=10, al='left')
    for c in (1, 3, 4, 5, 6, 7, 8):
        put(ws, 7, c, None, sz=10)

    r = 8
    for d in m['rows']:
        put(ws, r, 1, d['no'])
        put(ws, r, 2, f"{d['tag']}  -  {d['desc']}".strip(' -'), al='left')
        put(ws, r, 3, d['watt'])
        put(ws, r, 4, '3 PH' if d['phase'] == 3 else '1 PH')
        if d['phase'] == 3:
            for col in ('E', 'F', 'G'):
                put(ws, r, {'E': 5, 'F': 6, 'G': 7}[col], f'=$C{r}/3', fmt=N1)
            put(ws, r, 8, f'=SUM(E{r}:G{r})/({s["voltage_ll"]}*1.732*{s["pf"]})', fmt=N1)
        else:
            for col, ci in (('E', 5), ('F', 6), ('G', 7)):
                put(ws, r, ci, f'=$C{r}' if d['ph'] == {'E': 'R', 'F': 'S', 'G': 'T'}[col] else None, fmt=N1)
            put(ws, r, 8, f'=SUM(E{r}:G{r})/{s["voltage_ln"]}', fmt=N1)
        r += 1

    last, gap = r - 1, r + 1
    tot = gap + 1
    put(ws, gap, 2, 'TOTAL LOAD (WATT)', al='left')
    put(ws, gap, 3, f'=SUM(C8:C{gap-1})', fmt=N1)
    put(ws, tot, 2, 'TOTAL PER PHASE (WATT)', al='left')
    for ci, L in ((5, 'E'), (6, 'F'), (7, 'G')):
        put(ws, tot, ci, f'=SUM({L}8:{L}{gap})', fmt=N1)
    put(ws, tot, 8, f'=SUM(H8:H{gap})', fmt=N1)

    put(ws, tot + 1, 2, f'ARUS PER PHASE (AMPERE) @ {s["voltage_ln"]}V', al='left')
    for ci, L in ((5, 'E'), (6, 'F'), (7, 'G')):
        put(ws, tot + 1, ci, f'={L}{tot}/{s["voltage_ln"]}', fmt=N1)
    put(ws, tot + 1, 8, 'IMBALANCE', al='left')
    put(ws, tot + 1, 9, f'=(MAX(E{tot}:G{tot})-MIN(E{tot}:G{tot}))/AVERAGE(E{tot}:G{tot})', fmt='0.00%')
    put(ws, tot + 1, 10, 'max 10%', al='left')

    put(ws, tot + 2, 2, 'TOTAL CONNECTED LOAD (WATT)', al='left')
    put(ws, tot + 2, 6, f'=E{tot}+F{tot}+G{tot}', fmt=N1)
    put(ws, tot + 2, 8, 'SPARE LOAD', al='left')
    put(ws, tot + 2, 9, s['spare_percent'] / 100.0, fmt='0%')
    put(ws, tot + 2, 10, f'=F{tot+2}*(1+I{tot+2})', fmt=N2)
    put(ws, tot + 2, 11, 'watt', al='left')

    put(ws, tot + 3, 2, f'TOTAL APPARENT POWER (VA) @ PF {s["pf"]}', al='left')
    put(ws, tot + 3, 6, f'=F{tot+2}/{s["pf"]}', fmt=N1)
    put(ws, tot + 3, 10, f'=J{tot+2}-F{tot+2}', fmt=N2)

    put(ws, tot + 4, 2, f'TOTAL CURRENT (AMPERE) 3 PHASE {s["voltage_ll"]}V', al='left')
    put(ws, tot + 4, 6, f'=F{tot+3}/(1.732*{s["voltage_ll"]})', fmt=N1)

    put(ws, tot + 5, 2, 'DESIGN CURRENT x 1.1 (AMPERE)', al='left')
    put(ws, tot + 5, 6, f'=F{tot+4}*1.1', fmt=N2)

    put(ws, tot + 8, 2, 'MCCB / INCOMING BREAKER SIZE (A)', al='left')
    put(ws, tot + 8, 6, m['incoming'], bold=True, fmt='0')

    for k, w in (('A', 6), ('B', 58), ('C', 14), ('D', 9), ('E', 14), ('F', 13),
                 ('G', 13), ('H', 12), ('I', 12), ('J', 14), ('K', 8)):
        ws.column_dimensions[k].width = w
    for rr in range(3, tot + 10):
        ws.row_dimensions[rr].height = 15
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.print_area = f'A1:K{tot+9}'
    return ws


def sheet_cable(wb, m):
    s = m['sys']
    ws = wb.create_sheet('CABLE SIZING')
    put(ws, 1, 1, f"PERHITUNGAN KABEL FEEDER  -  {m['panel'].get('name','PANEL')}",
        bold=True, sz=12, al='left', border=False)
    put(ws, 2, 1, 'Basis: IEC 60364-5-52 (ampacity & faktor koreksi) + IEC 60364-5-54 (konduktor proteksi)',
        sz=8, al='left', border=False)
    put(ws, 4, 1, 'DATA DESAIN  (sel kuning = boleh diubah)', bold=True, sz=10, fill=HDR, al='left')
    ws.merge_cells('A4:D4')
    rows = [('Ib  - Arus beban desain', round(m['design'], 2), 'A', N2),
            ('In  - Rating MCCB terpasang', m['incoming'], 'A', '0'),
            ('U   - Tegangan sistem (line-line)', s['voltage_ll'], 'V', '0'),
            ('cos phi  - Faktor daya', s['pf'], '-', '0.00'),
            ('L   - Panjang kabel feeder', s['cable_length_m'], 'm', '0'),
            ('Suhu ambient rencana', s['ambient_c'], 'deg C', '0'),
            ('Ca  - Faktor koreksi suhu (XLPE)', s['ca'], '-', '0.00'),
            ('Batas drop tegangan', s['vd_limit'], '-', '0%')]
    r = 5
    for name, val, unit, fmt in rows:
        put(ws, r, 1, name, al='left')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        put(ws, r, 3, val, bold=True, fill=INP, fmt=fmt)
        put(ws, r, 4, unit, sz=8)
        r += 1
    IB, IN_, U, PF, L, CA, VD = 'C5', 'C6', 'C7', 'C8', 'C9', 'C11', 'C12'

    top = 15
    put(ws, top, 1, 'ALTERNATIF KABEL FEEDER', bold=True, sz=10, fill=HDR, al='left')
    ws.merge_cells(start_row=top, start_column=1, end_row=top, end_column=12)
    H = ['OPSI', 'KONFIGURASI KABEL', 'JML\nRUN', 'LUAS\nmm2', 'Iz TABEL\nA/run (30C)', 'Ca', 'Cg',
         'Iz TOTAL\nTERKOREKSI A', 'Iz >= In ?', 'DROP\nTEGANGAN V', 'DROP\nTEGANGAN %', 'HASIL']
    hr = top + 1
    for i, h in enumerate(H, 1):
        put(ws, hr, i, h, bold=True, sz=8, fill=HDR, wrap=True)
    ws.row_dimensions[hr].height = 42

    from panel_core import IZ_CU_XLPE_E
    r = hr + 1
    for i, o in enumerate(m['cables']):
        put(ws, r, 1, chr(65 + i), bold=True)
        put(ws, r, 2, o['desc'], al='left')
        put(ws, r, 3, o['runs'], fmt='0')
        put(ws, r, 4, o['mm2'] or '-', fmt='0')
        put(ws, r, 5, round(IZ_CU_XLPE_E.get(o['mm2'], o['iz']), 0), fill=INP, fmt='0')
        put(ws, r, 6, f'={CA}', fmt='0.00')
        put(ws, r, 7, 1.00 if o['runs'] == 1 else 0.88, fill=INP, fmt='0.00')
        put(ws, r, 8, f'=E{r}*F{r}*G{r}*C{r}', bold=True, fmt=N1)
        put(ws, r, 9, f'=IF(H{r}>={IN_},"OK","TIDAK")', bold=True)
        put(ws, r, 10, round(o['vd'], 2), fmt=N2)
        put(ws, r, 11, f'=J{r}/{U}', fmt='0.00%')
        put(ws, r, 12, f'=IF(AND(I{r}="OK",K{r}<={VD}),"MEMENUHI","TIDAK MEMENUHI")', bold=True)
        r += 1

    r += 1
    put(ws, r, 1, 'CATATAN', bold=True, sz=10, fill=HDR, al='left')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    notes = [
        '1.  Iz TABEL (sel kuning) = IEC 60364-5-52 Cu/XLPE 90C, 3 konduktor berbeban, 30C, metode E (kabel multicore di cable tray).',
        '     Nilai ini konservatif; datasheet pabrikan umumnya lebih tinggi. Ganti sel kuning dengan data pabrikan, semua hasil terhitung ulang.',
        f'2.  Ca mengikuti ambient. 30C=1,00  35C=0,96  40C=0,91  45C=0,87.  Nilai terpakai: {s["ca"]}.',
        '3.  Cg = faktor pengelompokan. 1 sirkit di tray = 1,00; 2 sirkit bersentuhan = 0,88; 3 sirkit = 0,82.',
        '4.  Syarat IEC 60364-4-43:  Ib <= In <= Iz.',
        '5.  Drop tegangan: dV = akar(3) x Ib x L x (R x cos phi + X x sin phi) / jumlah run.',
        '6.  Opsi paralel: panjang, penampang, bahan dan cara pemasangan tiap run wajib identik (IEC 60364-5-52 cl.523.7).',
        f'7.  Konduktor proteksi (PE) mengikuti IEC 60364-5-54: S > 35 mm2 -> PE = S/2.',
    ]
    for n in notes:
        put(ws, r, 1, n, sz=8, al='left', border=False)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        r += 1

    for k, w in (('A', 7), ('B', 46), ('C', 7), ('D', 8), ('E', 12), ('F', 7), ('G', 7),
                 ('H', 14), ('I', 11), ('J', 12), ('K', 12), ('L', 17)):
        ws.column_dimensions[k].width = w
    ws.freeze_panes = f'A{hr+1}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    return ws


def make(cfg, outdir='.'):
    """Write the workbook and return (path, summary line).

    Mirrors build_dxf.make() so a caller that has already parsed loads.json -
    the serverless function in api/build.py - can produce the file without
    going back through argv and a subprocess.
    """
    os.makedirs(outdir, exist_ok=True)
    m = build(cfg)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_schedule(wb, m)
    sheet_cable(wb, m)
    path = os.path.join(outdir, f"PANEL_SCHEDULE_{safe_name(m['panel'].get('name','PANEL'))}.xlsx")
    wb.save(path)
    summary = (f"circuits={len(m['rows'])}  total={m['total_w']:,.0f} W  "
               f"R/S/T={m['phase_w']['R']:,.0f}/{m['phase_w']['S']:,.0f}/{m['phase_w']['T']:,.0f} W  "
               f"imbalance={m['imbalance']*100:.2f}%  Ib={m['design']:.1f} A  "
               f"incoming={m['incoming']} A")
    return path, summary


def main():
    # See build_dxf.main(): a rejected loads.json should read as a message, not
    # as a traceback with the message somewhere in the middle of it.
    try:
        cfg = load_config(sys.argv[1])
        outdir = sys.argv[2] if len(sys.argv) > 2 else '.'
        path, summary = make(cfg, outdir)
    except ValueError as exc:
        print(exc, file=sys.stderr)
        sys.exit(2)
    print(f"OK  {path}")
    print(f"    {summary}")


if __name__ == '__main__':
    main()
